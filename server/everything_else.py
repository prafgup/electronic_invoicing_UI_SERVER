import argparse

from pytesseract import Output
import pytesseract
import cv2
from openpyxl import load_workbook
import csv
import json
import string

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--inputPath', metavar='', required=True, help="Image File Path Here")
    parser.add_argument('-j', '--jsonPath', metavar='', required=True, help="Json File Path Here")
    parser.add_argument('-w', '--worksheetPath', metavar='', required=True, help="Excel Worksheet Path Here")
    args = vars(parser.parse_args())
    filename = str(args["inputPath"])
    jsonPath = str(args["jsonPath"])
    sheetPath = str(args["worksheetPath"])
    # load the input image, convert it from BGR to RGB channel ordering,
    # and use Tesseract to localize each area of text in the input image
    image = cv2.imread(r'./../web/invoice_template/src/preprocessed/' + filename)
    rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    # tessdata_dir_config = '"--tessdata-dir "D:\Teserract-OCR\tesdata"'
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    results = pytesseract.image_to_data(rgb, output_type=Output.DICT, timeout=30)
    boxes = []
    # loop over each of the individual text localizations
    for i in range(0, len(results["text"])):
        # extract the bounding box coordinates of the text region from
        # the current result
        x = results["left"][i]
        y = results["top"][i]
        w = results["width"][i]
        h = results["height"][i]

        # extract the OCR text itself along with the confidence of the
        # text localization
        text = results["text"][i]
        conf = int(results["conf"][i])

        # filter out weak confidence text localizations
        if conf > 0:
            # display the confidence and text to our terminal
            # print("Confidence: {}".format(conf))
            # print("Text: {}".format(text))

            temp = [text, [x - 5, y - 5, w + 5, h + 5], conf]
            # print(temp)
            # print("")
            if (len(text) > 2):
                boxes.append(temp)

            # strip out non-ASCII text so we can draw the text on the image
            # using OpenCV, then draw a bounding box around the text along
            # with the text itself
            text = "".join([c if ord(c) < 128 else "" for c in text]).strip()
            # cv2.rectangle(image, (x, y), (x + w, y + h), (0, 255, 0), 2)
            # cv2.putText(image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX,
            #             1.2, (0, 0, 255), 1)

    # show the output image
    # cv2.imshow("abcd",image)
    # cv2.waitKey(0)

    jsonData = ""
    # Read Data from json file here. Assuming a format of type data2= [[code,[x,y,w,h]],[code,[x,y,w,h]],....]
    with open(r'./../database/templates.json') as BOB:
        jsonData = json.load(BOB)
    parsedData = []

    hhw = 1080
    hhh  =1920


    for x in jsonData:
        if x['name'] == jsonPath:
            y = x['data_points']
            # print(y)
            for elem in y:
                parsedData.append([int(elem['cls']), [int(float(elem['x']) * hhw), int(float(elem['y']) * hhh),
                                                      int(float(elem['w']) * hhw), int(float(elem['h']) * hhh)]])
    #print(parsedData)
    lastX = 1000
    lastY = 1900
    lastW = 0
    lastH = 0

    minYT = 1000000
    minXT = 1000000
    minYD = 1000000
    minXD = 1000000
    maxXD = 0
    maxYD = 0
    for row in parsedData:
        if row[1][1] < minYT:
            minYT = row[1][1]
            minXT = row[1][0]

    for row in boxes:
        if row[1][1] < minYD:
            minYD = row[1][1]
            minXD = row[1][0]
        if row[1][1] > maxYD:
            maxYD = row[1][1]
            maxXD = row[1][0]

    diffX = minXT - minXD
    diffY = minYT - minYD

    for row in parsedData:
        row[1][0] += diffX
        row[1][1] += diffY

    for elem in parsedData:
        if elem[0] == -1:
            lastX = elem[1][0]
            lastY = elem[1][1]
            lastW = elem[1][2]
            lastH = elem[1][3]

    diff =  (lastY - maxYD )*-1

    for elem in parsedData:
        if int(elem[0]) >= 21:
            elem[1][3] += diff
    #print(lastY , maxYD )
    #print(parsedData)

    image = cv2.imread(r'./../web/invoice_template/src/preprocessed/' + filename)
    errorImage = image
    # data2=[[1,[160,427,308,35]]]
    wb = load_workbook(filename="sheet.xlsx")
    ws = wb.active
    errorPresent=False
    for row in parsedData:
        x = row[1][0]
        y = row[1][1]
        w = row[1][2]
        h = row[1][3]
        crop = image[y - 5:y + h + 5, x - 5:x + w + 5]
        # cv2_imshow(crop)
        crop_rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        text = pytesseract.image_to_string(crop_rgb, lang='eng', timeout=10)
        crop_data=pytesseract.image_to_data(crop,lang='eng',output_type=Output.DICT)
        av_accuracy=0
        counter=0
        for i in range(len(crop_data['text'])):
            if len(crop_data['text'][i])>2:
                av_accuracy+=int(crop_data['conf'][i])
                counter+=1
        if counter>0:
            av_accuracy/=counter

        #print(av_accuracy)
        if av_accuracy<50 and av_accuracy!=0:
            cv2.rectangle(errorImage, (x-5, y-5), (x + w+5, y + h+5), (0, 255, 0), 2)
            errorPresent=True
        # print(text)
        text.encode("ascii", "ignore")
        text = "".join([c if ord(c) < 128 else "" for c in text]).strip()
        text = str(text)
        code = row[0]
        if code == 1:
            # print("Hello")

            ws.cell(row=3, column=5).value = str(text).encode('ascii', errors='ignore')
        # print(ws.cell(row=3,column=5).value)

        if code == 2:
            ws.cell(row=4, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 3:
            ws.cell(row=5, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 4:
            ws.cell(row=6, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 5:
            ws.cell(row=11, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 6:
            ws.cell(row=12, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 7:
            ws.cell(row=13, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 8:
            ws.cell(row=14, column=5).value = str(text).encode('ascii', errors='ignore')

        if code == 9:
            ws.cell(row=3, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 10:
            ws.cell(row=4, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 11:
            ws.cell(row=5, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 12:
            ws.cell(row=6, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 13:
            ws.cell(row=7, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 14:
            ws.cell(row=8, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 15:
            ws.cell(row=9, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 16:
            ws.cell(row=10, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 17:
            ws.cell(row=11, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 18:
            ws.cell(row=12, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 19:
            ws.cell(row=13, column=13).value = str(text).encode('ascii', errors='ignore')

        if code == 20:
            ws.cell(row=14, column=13).value = str(text).encode('ascii', errors='ignore')

        if code >= 21:
            lines = text.splitlines()
            for i in range(len(lines)):
                ws.cell(row=18 + i, column=2 + code - 21).value = str(lines[i]).encode('ascii', errors='ignore')
    wb.save(sheetPath)
    cv2.imwrite(r'./../web/invoice_template/src/preprocessed/errorImage.png',errorImage)

    print(errorPresent)
    wb.close()
